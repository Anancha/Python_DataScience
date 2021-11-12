import openpyxl as xl
import numpy as np

wb = xl.Workbook()
ws = wb.worksheets[0]

data = np.random.randint(1000,9999,10)
row = 1

for i in data:
    ws.cell(row,1).value = i

    if i>=5000:
        ws.cell(row,2).value = 'ค่ามาก'
    else:
        ws.cell(row,2).value = 'ค่าน้อย'

    row = row+1
wb.save('data.xlsx')
print('สร้างไฟล์ Excel เสร็จเรียบร้อยแล้ว')

